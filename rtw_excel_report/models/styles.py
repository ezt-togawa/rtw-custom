import logging

from odoo import api, models

_logger = logging.getLogger(__name__)

try:
    from openpyxl.styles import Alignment, Font, PatternFill , Border , Side
except ImportError:
    _logger.debug('Cannot import "openpyxl". Please make sure it is installed.')

class XLSXStyles(models.AbstractModel):
    _inherit = "xlsx.styles"
    _description = "Available styles for excel"

    @api.model
    def get_openpyxl_styles(self):
        return {
            "font": {
                "bold": Font(name="Arial", size=10, bold=True),
                "bold_red": Font(name="Arial", size=10, color="FF0000", bold=True),
            },
            "fill": {
                "red": PatternFill("solid", fgColor="FF0000"),
                "grey": PatternFill("solid", fgColor="DDDDDD"),
                "light_grey": PatternFill("solid", fgColor="F5F5F5"),
                "yellow": PatternFill("solid", fgColor="FFFCB7"),
                "blue": PatternFill("solid", fgColor="9BF3FF"),
                "green": PatternFill("solid", fgColor="B0FF99"),
                "table_cell": PatternFill("solid", fgColor="F2F2F2"),
            },
            "align": {
                "left": Alignment(horizontal="left"),
                "center": Alignment(horizontal="center"),
                "right": Alignment(horizontal="right"),
            },
            "style": {
                "number": "#,##0.00",
                "date": "dd/mm/yyyy",
                "datestamp": "yyyy-mm-dd",
                "text": "@",
                "percent": "0.00%",
            },
            "border": {
                "thin": Border(
                    left=Side(style="thin", color="000000" ),
                    right=Side(style="thin", color="333333"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000")
                ),
                "medium": Border(
                    left=Side(style="medium", color="000000"),
                    right=Side(style="medium", color="333333"),
                    top=Side(style="medium", color="000000"),
                    bottom=Side(style="medium", color="000000")
                ),
            },
        }
